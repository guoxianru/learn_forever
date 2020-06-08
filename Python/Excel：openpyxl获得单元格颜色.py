# -*- coding: utf-8 -*-
# -*- author: GXR -*-

"""
    获取已有工作表的颜色
    pip install openpyxl
"""

from openpyxl import load_workbook

# 打开工作簿
workbook = load_workbook(r"test.xls")
# 打开工作表
worksheet = workbook.get_sheet_by_name("Sheet1")
# 最大行列数
rows, cols = worksheet.max_row, worksheet.max_column
# 循坏获取单元格
for i in range(1, rows + 1):
    for j in range(1, cols + 1):
        # 获取单元格
        ce = worksheet.cell(row=i, column=j)
        fill = ce.fill
        font = ce.font
        # 单元格颜色
        if font.color.rgb == "FF00B050":
            print(ce.value)
